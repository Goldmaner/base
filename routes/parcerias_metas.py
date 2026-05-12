"""
Blueprint de Quadro de Metas — Plano de Trabalho
Hierarquia: SEI → Objetivo (indicadores/meios compartilhados) → Metas individuais
"""

import csv
import io
import math
from datetime import datetime
from flask import (
    Blueprint, render_template, request, jsonify, session, Response
)
from db import get_cursor, get_db
from utils import login_required
from decorators import requires_access

parcerias_metas_bp = Blueprint(
    'parcerias_metas', __name__, url_prefix='/parcerias-metas'
)


# ── Helpers de retroalimentação ───────────────────────────────────────────────

def _resolve_indicador(texto, usuario):
    """Retorna id do indicador pelo texto (str), criando-o se não existir."""
    if not texto:
        return None
    cur = get_cursor()
    cur.execute(
        "SELECT id FROM categoricas.c_dgp_indicadores WHERE LOWER(indicador) = LOWER(%s)",
        (texto,)
    )
    row = cur.fetchone()
    if row:
        return row['id']
    cur.execute(
        "INSERT INTO categoricas.c_dgp_indicadores (indicador, criado_por) "
        "VALUES (%s, %s) RETURNING id",
        (texto, usuario)
    )
    return cur.fetchone()['id']


def _resolve_indicadores(textos, usuario):
    """Resolve lista de textos → lista de ids (cria os que não existem)."""
    ids = []
    for t in textos:
        t = (t or '').strip()
        if t:
            ids.append(_resolve_indicador(t, usuario))
    return ids or None


def _resolve_meio(texto, usuario):
    """Retorna id do meio de aferição pelo texto (str), criando-o se não existir."""
    if not texto:
        return None
    cur = get_cursor()
    cur.execute(
        "SELECT id FROM categoricas.c_dgp_meios_afericao WHERE LOWER(meios_afericao) = LOWER(%s)",
        (texto,)
    )
    row = cur.fetchone()
    if row:
        return row['id']
    cur.execute(
        "INSERT INTO categoricas.c_dgp_meios_afericao (meios_afericao, criado_por) "
        "VALUES (%s, %s) RETURNING id",
        (texto, usuario)
    )
    return cur.fetchone()['id']


def _resolve_meios(textos, usuario):
    """Resolve lista de textos → lista de ids (cria os que não existem)."""
    ids = []
    for t in textos:
        t = (t or '').strip()
        if t:
            ids.append(_resolve_meio(t, usuario))
    return ids or None


# ── Rota principal ────────────────────────────────────────────────────────────

@parcerias_metas_bp.route("/", methods=["GET"])
@login_required
@requires_access('parcerias_metas')
def index():
    cur = get_cursor()
    filtro_sei    = request.args.get('filtro_sei',    '').strip()
    filtro_ind    = request.args.get('filtro_ind',    '').strip()
    filtro_meio   = request.args.get('filtro_meio',   '').strip()
    filtro_texto  = request.args.get('filtro_texto',  '').strip()

    # Catálogos para modais
    cur.execute("""
        SELECT id, meta_tipo, tipo_classificacao
        FROM categoricas.c_dgp_meta_tipos
        ORDER BY tipo_classificacao NULLS LAST, meta_tipo
    """)
    meta_tipos = cur.fetchall()

    cur.execute("SELECT id, indicador FROM categoricas.c_dgp_indicadores ORDER BY indicador")
    indicadores_cat = cur.fetchall()

    cur.execute("SELECT id, meios_afericao FROM categoricas.c_dgp_meios_afericao ORDER BY meios_afericao")
    meios_cat = cur.fetchall()

    where_clause = ""
    conditions = []
    params = []
    if filtro_sei:
        conditions.append("co.sei_numero ILIKE %s")
        params.append(f"%{filtro_sei}%")
    if filtro_ind:
        conditions.append("""EXISTS (
            SELECT 1 FROM categoricas.c_dgp_indicadores _ind
            WHERE _ind.id = ANY(co.indicadores_ids)
              AND LOWER(_ind.indicador) = LOWER(%s)
        )""")
        params.append(filtro_ind)
    if filtro_meio:
        conditions.append("""EXISTS (
            SELECT 1 FROM categoricas.c_dgp_meios_afericao _ma
            WHERE _ma.id = ANY(co.meios_afericao_ids)
              AND LOWER(_ma.meios_afericao) = LOWER(%s)
        )""")
        params.append(filtro_meio)
    if filtro_texto:
        conditions.append("""(
            co.objetivo ILIKE %s
            OR EXISTS (
                SELECT 1 FROM celebracao.celebracao_metas _cm
                WHERE _cm.objetivo_id = co.id
                  AND (_cm.meta_titulo ILIKE %s OR _cm.meta_descricao ILIKE %s)
            )
        )""")
        params.extend([f"%{filtro_texto}%"] * 3)
    if conditions:
        where_clause = "WHERE " + " AND ".join(conditions)

    cur.execute(f"""
        SELECT
            co.id                                              AS obj_id,
            co.sei_numero,
            co.objetivo,
            co.indicadores_ids,
            co.indicadores_ni,
            co.meta_obs_indicadores,
            co.meios_afericao_ids,
            co.meios_ni,
            co.ordem                                           AS obj_ordem,
            co.criado_por                                      AS obj_criado_por,
            co.criado_em                                       AS obj_criado_em,
            COALESCE(
                (SELECT STRING_AGG(ind.indicador, ' | ' ORDER BY ind.indicador)
                 FROM categoricas.c_dgp_indicadores ind
                 WHERE ind.id = ANY(co.indicadores_ids)),
                NULL
            ) AS indicador_label,
            COALESCE(
                (SELECT STRING_AGG(ma.meios_afericao, ' | ' ORDER BY ma.meios_afericao)
                 FROM categoricas.c_dgp_meios_afericao ma
                 WHERE ma.id = ANY(co.meios_afericao_ids)),
                NULL
            ) AS meios_label,
            COALESCE(p.numero_termo, cp.numero_termo, '-')     AS numero_termo,
            COALESCE(p.osc, cp.osc, '-')                       AS osc,
            cp.status                                          AS status,
            cp.substatus                                       AS substatus,
            cm.id                                              AS meta_id,
            cm.meta_titulo,
            cm.meta_descricao,
            cm.meta_tipo_ids,
            cm.tipos_ni,
            cm.observacoes,
            cm.ordem                                           AS meta_ordem,
            cm.criado_por                                      AS meta_criado_por,
            cm.criado_em                                       AS meta_criado_em,
            cm.atualizado_por                                  AS meta_atualizado_por,
            cm.atualizado_em                                   AS meta_atualizado_em
        FROM celebracao.celebracao_objetivos co
        LEFT JOIN celebracao.celebracao_metas cm
            ON cm.objetivo_id = co.id
        LEFT JOIN (SELECT sei_celeb, MAX(numero_termo) AS numero_termo, MAX(osc) AS osc
                   FROM public.parcerias GROUP BY sei_celeb) p
            ON p.sei_celeb = co.sei_numero
        LEFT JOIN (SELECT sei_celeb, MAX(numero_termo) AS numero_termo, MAX(osc) AS osc, MAX(status) AS status, MAX(substatus) AS substatus
                   FROM celebracao.celebracao_parcerias GROUP BY sei_celeb) cp
            ON cp.sei_celeb = co.sei_numero
        {where_clause}
        ORDER BY co.sei_numero, co.ordem, co.id, cm.ordem, cm.id
    """, params)
    rows = cur.fetchall()

    tipo_map = {r['id']: r              for r in meta_tipos}
    ind_map  = {r['id']: r['indicador'] for r in indicadores_cat}
    meio_map = {r['id']: r['meios_afericao'] for r in meios_cat}

    # Agrupa: sei_groups[sei] → {numero_termo, osc, objetivos: [{obj, metas: [...]}]}
    sei_groups = {}
    obj_cache  = {}

    for r in rows:
        sei    = r['sei_numero']
        obj_id = r['obj_id']

        if sei not in sei_groups:
            sei_groups[sei] = {
                'sei_numero':   sei,
                'numero_termo': r['numero_termo'],
                'osc':          r['osc'],
                'status':       r['status'],
                'substatus':    r['substatus'],
                'objetivos':    [],
            }

        if obj_id not in obj_cache:
            ind_ids  = list(r['indicadores_ids']   or [])
            meio_ids = list(r['meios_afericao_ids'] or [])
            obj = {
                'id':                 obj_id,
                'sei_numero':         sei,
                'objetivo':           r['objetivo'],
                'indicadores_ni':     r['indicadores_ni'],
                'meios_ni':           r['meios_ni'],
                'indicador_label':    r['indicador_label'],
                'meios_label':        r['meios_label'],
                'indicadores_textos': [ind_map[i] for i in ind_ids  if i in ind_map],
                'meios_textos':       [meio_map[i] for i in meio_ids if i in meio_map],
                'obs_indicadores':    list(r['meta_obs_indicadores'] or []),
                'obj_ordem':          r['obj_ordem'],
                'criado_por':         r['obj_criado_por'],
                'criado_em':          r['obj_criado_em'],
                'metas':              [],
            }
            sei_groups[sei]['objetivos'].append(obj)
            obj_cache[obj_id] = obj

        if r['meta_id'] is not None:
            raw_tipos = list(r['meta_tipo_ids'] or [])
            obj_cache[obj_id]['metas'].append({
                'id':                r['meta_id'],
                'meta_titulo':       r['meta_titulo'],
                'meta_descricao':    r['meta_descricao'],
                'meta_tipo_ids':     raw_tipos,
                'tipos_ni':          r['tipos_ni'],
                'meta_tipos_labels': [tipo_map[i] for i in raw_tipos if i in tipo_map],
                'observacoes':       r['observacoes'],
                'meta_ordem':        r['meta_ordem'],
                'criado_por':        r['meta_criado_por'],
                'criado_em':         r['meta_criado_em'],
                'atualizado_por':    r['meta_atualizado_por'],
                'atualizado_em':     r['meta_atualizado_em'],
            })

    sei_list    = list(sei_groups.values())
    total_seis  = len(sei_list)
    per_page    = 50
    total_pages = max(1, math.ceil(total_seis / per_page))
    page        = min(max(1, int(request.args.get('page', 1) or 1)), total_pages)
    sei_page    = sei_list[(page - 1) * per_page : page * per_page]

    return render_template(
        'parcerias_metas.html',
        sei_groups=sei_page,
        meta_tipos=meta_tipos,
        filtro_sei=filtro_sei,
        filtro_ind=filtro_ind,
        filtro_meio=filtro_meio,
        filtro_texto=filtro_texto,
        page=page,
        total_pages=total_pages,
        total_seis=total_seis,
    )



# ── Dicionário de Indicadores e Meios de Aferição ─────────────────────────────

@parcerias_metas_bp.route("/dicionario", methods=["GET"])
@login_required
@requires_access('parcerias_metas')
def dicionario():
    cur = get_cursor()
    cur.execute("""
        SELECT t.id, t.indicador, t.descricao, t.observacao,
               COUNT(co.id) AS qtd_usos
        FROM categoricas.c_dgp_indicadores t
        LEFT JOIN celebracao.celebracao_objetivos co ON t.id = ANY(co.indicadores_ids)
        GROUP BY t.id, t.indicador, t.descricao, t.observacao
        ORDER BY t.indicador
    """)
    indicadores = cur.fetchall()

    cur.execute("""
        SELECT t.id, t.meios_afericao, t.descricao, t.observacao,
               COUNT(co.id) AS qtd_usos
        FROM categoricas.c_dgp_meios_afericao t
        LEFT JOIN celebracao.celebracao_objetivos co ON t.id = ANY(co.meios_afericao_ids)
        GROUP BY t.id, t.meios_afericao, t.descricao, t.observacao
        ORDER BY t.meios_afericao
    """)
    meios_afericao = cur.fetchall()

    return render_template(
        'parcerias_metas_dicionario.html',
        indicadores=indicadores,
        meios_afericao=meios_afericao,
    )


# ── API: SEI numbers (3 fontes) ───────────────────────────────────────────────

@parcerias_metas_bp.route("/api/sei-numeros", methods=["GET"])
@login_required
def api_sei_numeros():
    cur = get_cursor()
    # SEIs firmados em parcerias (prioridade máxima — sem rótulo de celebração)
    cur.execute("""
        WITH sei_parceria AS (
            SELECT TRIM(sei_celeb) AS sei_numero
            FROM public.parcerias
            WHERE sei_celeb IS NOT NULL AND TRIM(sei_celeb) != ''
        ),
        sei_celebracao AS (
            SELECT TRIM(sei_celeb) AS sei_numero
            FROM celebracao.celebracao_parcerias
            WHERE sei_celeb IS NOT NULL AND TRIM(sei_celeb) != ''
        ),
        sei_edital AS (
            SELECT TRIM(edital_processo_sei) AS sei_numero
            FROM public.parcerias_edital
            WHERE edital_processo_sei IS NOT NULL AND TRIM(edital_processo_sei) != ''
        )
        SELECT sei_numero, 'Parceria' AS fonte FROM sei_parceria
        UNION
        -- Celebrações só aparecem se não há parceria firmada com o mesmo SEI
        SELECT sei_numero, 'Celebração' AS fonte
        FROM sei_celebracao
        WHERE sei_numero NOT IN (SELECT sei_numero FROM sei_parceria)
        UNION
        SELECT sei_numero, 'Edital' AS fonte
        FROM sei_edital
        WHERE sei_numero NOT IN (SELECT sei_numero FROM sei_parceria)
          AND sei_numero NOT IN (SELECT sei_numero FROM sei_celebracao)
        ORDER BY sei_numero
    """)
    rows = cur.fetchall()
    return jsonify([{'sei_numero': r['sei_numero'], 'fonte': r['fonte']} for r in rows])


# ── API: Meta tipos agrupados ─────────────────────────────────────────────────

@parcerias_metas_bp.route("/api/meta-tipos", methods=["GET"])
@login_required
def api_meta_tipos():
    cur = get_cursor()
    cur.execute("""
        SELECT id, meta_tipo, tipo_classificacao, descricao
        FROM categoricas.c_dgp_meta_tipos
        ORDER BY tipo_classificacao NULLS LAST, meta_tipo
    """)
    rows = cur.fetchall()

    # Agrupar por tipo_classificacao
    grupos = {}
    for r in rows:
        cls = r['tipo_classificacao'] or 'Sem classificação'
        if cls not in grupos:
            grupos[cls] = []
        grupos[cls].append({'id': r['id'], 'meta_tipo': r['meta_tipo'], 'descricao': r['descricao']})

    return jsonify(grupos)


# ── API: Definições (glossário) ───────────────────────────────────────────────

@parcerias_metas_bp.route("/api/definicoes", methods=["GET"])
@login_required
def api_definicoes():
    cur = get_cursor()
    cur.execute("""
        SELECT id, conceito, definicao, observacoes
        FROM categoricas.c_dgp_plano_definicoes
        ORDER BY conceito
    """)
    rows = cur.fetchall()
    return jsonify([dict(r) for r in rows])


# ── API: Indicadores (retroalimentação + dicionário) ──────────────────────────

@parcerias_metas_bp.route("/api/indicadores", methods=["GET"])
@login_required
def api_indicadores():
    cur = get_cursor()
    cur.execute("SELECT id, indicador, descricao, observacao FROM categoricas.c_dgp_indicadores ORDER BY indicador")
    return jsonify([dict(r) for r in cur.fetchall()])


@parcerias_metas_bp.route("/api/indicadores", methods=["POST"])
@login_required
@requires_access('parcerias_metas')
def api_indicador_criar():
    data = request.get_json() or {}
    usuario = session.get('d_usuario', 'sistema')
    indicador = (data.get('indicador') or '').strip().rstrip('.;,')
    if not indicador:
        return jsonify({'erro': 'indicador é obrigatório'}), 400
    cur = get_cursor()
    # Verifica duplicata antes de inserir
    cur.execute("SELECT id FROM categoricas.c_dgp_indicadores WHERE LOWER(indicador) = LOWER(%s)", (indicador,))
    row = cur.fetchone()
    if row:
        return jsonify({'erro': 'Indicador já existe', 'id': row['id']}), 409
    cur.execute(
        "INSERT INTO categoricas.c_dgp_indicadores (indicador, descricao, observacao, criado_por) "
        "VALUES (%s, %s, %s, %s) RETURNING id",
        (indicador, data.get('descricao') or None, data.get('observacao') or None, usuario)
    )
    novo_id = cur.fetchone()['id']
    get_db().commit()
    return jsonify({'sucesso': True, 'id': novo_id}), 201


@parcerias_metas_bp.route("/api/indicadores/<int:ind_id>", methods=["PUT"])
@login_required
@requires_access('parcerias_metas')
def api_indicador_editar(ind_id):
    data = request.get_json() or {}
    indicador = (data.get('indicador') or '').strip().rstrip('.;,')
    if not indicador:
        return jsonify({'erro': 'indicador é obrigatório'}), 400
    cur = get_cursor()
    cur.execute("""
        UPDATE categoricas.c_dgp_indicadores
        SET indicador = %s, descricao = %s, observacao = %s
        WHERE id = %s
    """, (indicador, data.get('descricao') or None, data.get('observacao') or None, ind_id))
    get_db().commit()
    return jsonify({'sucesso': True})


@parcerias_metas_bp.route("/api/indicadores/<int:ind_id>", methods=["DELETE"])
@login_required
@requires_access('parcerias_metas')
def api_indicador_excluir(ind_id):
    cur = get_cursor()
    cur.execute("DELETE FROM categoricas.c_dgp_indicadores WHERE id = %s", (ind_id,))
    get_db().commit()
    return jsonify({'sucesso': True})


@parcerias_metas_bp.route("/api/indicadores/<int:ind_id>/usos", methods=["GET"])
@login_required
def api_indicador_usos(ind_id):
    """Retorna lista de objetivos que usam este indicador."""
    cur = get_cursor()
    cur.execute("""
        SELECT co.sei_numero, co.objetivo,
               COALESCE(p.numero_termo, cp.numero_termo, '-') AS numero_termo,
               COALESCE(p.osc, cp.osc, '-') AS osc
        FROM celebracao.celebracao_objetivos co
        LEFT JOIN (SELECT sei_celeb, MAX(numero_termo) AS numero_termo, MAX(osc) AS osc
                   FROM public.parcerias GROUP BY sei_celeb) p ON p.sei_celeb = co.sei_numero
        LEFT JOIN (SELECT sei_celeb, MAX(numero_termo) AS numero_termo, MAX(osc) AS osc
                   FROM celebracao.celebracao_parcerias GROUP BY sei_celeb) cp ON cp.sei_celeb = co.sei_numero
        WHERE %s = ANY(co.indicadores_ids)
        ORDER BY co.sei_numero, co.objetivo
    """, (ind_id,))
    return jsonify([dict(r) for r in cur.fetchall()])


# ── API: Meios de Aferição (retroalimentação + dicionário) ────────────────────

@parcerias_metas_bp.route("/api/meios-afericao", methods=["GET"])
@login_required
def api_meios():
    cur = get_cursor()
    cur.execute("SELECT id, meios_afericao, descricao FROM categoricas.c_dgp_meios_afericao ORDER BY meios_afericao")
    return jsonify([dict(r) for r in cur.fetchall()])


@parcerias_metas_bp.route("/api/meios-afericao", methods=["POST"])
@login_required
@requires_access('parcerias_metas')
def api_meio_criar():
    data = request.get_json() or {}
    usuario = session.get('d_usuario', 'sistema')
    meios_afericao = (data.get('meios_afericao') or '').strip().rstrip('.;,')
    if not meios_afericao:
        return jsonify({'erro': 'meios_afericao é obrigatório'}), 400
    cur = get_cursor()
    cur.execute("SELECT id FROM categoricas.c_dgp_meios_afericao WHERE LOWER(meios_afericao) = LOWER(%s)", (meios_afericao,))
    row = cur.fetchone()
    if row:
        return jsonify({'erro': 'Meio de aferição já existe', 'id': row['id']}), 409
    cur.execute(
        "INSERT INTO categoricas.c_dgp_meios_afericao (meios_afericao, descricao, observacao, criado_por) "
        "VALUES (%s, %s, %s, %s) RETURNING id",
        (meios_afericao, data.get('descricao') or None, data.get('observacao') or None, usuario)
    )
    novo_id = cur.fetchone()['id']
    get_db().commit()
    return jsonify({'sucesso': True, 'id': novo_id}), 201


@parcerias_metas_bp.route("/api/meios-afericao/<int:meio_id>", methods=["PUT"])
@login_required
@requires_access('parcerias_metas')
def api_meio_editar(meio_id):
    data = request.get_json() or {}
    meios_afericao = (data.get('meios_afericao') or '').strip().rstrip('.;,')
    if not meios_afericao:
        return jsonify({'erro': 'meios_afericao é obrigatório'}), 400
    cur = get_cursor()
    cur.execute("""
        UPDATE categoricas.c_dgp_meios_afericao
        SET meios_afericao = %s, descricao = %s, observacao = %s
        WHERE id = %s
    """, (meios_afericao, data.get('descricao') or None, data.get('observacao') or None, meio_id))
    get_db().commit()
    return jsonify({'sucesso': True})


@parcerias_metas_bp.route("/api/meios-afericao/<int:meio_id>", methods=["DELETE"])
@login_required
@requires_access('parcerias_metas')
def api_meio_excluir(meio_id):
    cur = get_cursor()
    cur.execute("DELETE FROM categoricas.c_dgp_meios_afericao WHERE id = %s", (meio_id,))
    get_db().commit()
    return jsonify({'sucesso': True})


@parcerias_metas_bp.route("/api/meios-afericao/<int:meio_id>/usos", methods=["GET"])
@login_required
def api_meio_usos(meio_id):
    """Retorna lista de objetivos que usam este meio de aferição."""
    cur = get_cursor()
    cur.execute("""
        SELECT co.sei_numero, co.objetivo,
               COALESCE(p.numero_termo, cp.numero_termo, '-') AS numero_termo,
               COALESCE(p.osc, cp.osc, '-') AS osc
        FROM celebracao.celebracao_objetivos co
        LEFT JOIN (SELECT sei_celeb, MAX(numero_termo) AS numero_termo, MAX(osc) AS osc
                   FROM public.parcerias GROUP BY sei_celeb) p ON p.sei_celeb = co.sei_numero
        LEFT JOIN (SELECT sei_celeb, MAX(numero_termo) AS numero_termo, MAX(osc) AS osc
                   FROM celebracao.celebracao_parcerias GROUP BY sei_celeb) cp ON cp.sei_celeb = co.sei_numero
        WHERE %s = ANY(co.meios_afericao_ids)
        ORDER BY co.sei_numero, co.objetivo
    """, (meio_id,))
    return jsonify([dict(r) for r in cur.fetchall()])


# ── Similaridade ─────────────────────────────────────────────────────────────

@parcerias_metas_bp.route("/api/indicadores/similares", methods=["GET"])
@login_required
@requires_access('parcerias_metas')
def api_indicador_similares():
    texto = (request.args.get('q') or '').strip()
    if not texto:
        return jsonify([])
    cur = get_cursor()
    cur.execute("""
        SELECT id, indicador
        FROM categoricas.c_dgp_indicadores
        WHERE indicador ILIKE %s
          AND LOWER(indicador) != LOWER(%s)
        ORDER BY indicador
        LIMIT 10
    """, (f"%{texto}%", texto))
    return jsonify([dict(r) for r in cur.fetchall()])


@parcerias_metas_bp.route("/api/meios-afericao/similares", methods=["GET"])
@login_required
@requires_access('parcerias_metas')
def api_meio_similares():
    texto = (request.args.get('q') or '').strip()
    if not texto:
        return jsonify([])
    cur = get_cursor()
    cur.execute("""
        SELECT id, meios_afericao
        FROM categoricas.c_dgp_meios_afericao
        WHERE meios_afericao ILIKE %s
          AND LOWER(meios_afericao) != LOWER(%s)
        ORDER BY meios_afericao
        LIMIT 10
    """, (f"%{texto}%", texto))
    return jsonify([dict(r) for r in cur.fetchall()])


# ── Excluir órfãos em lote ────────────────────────────────────────────────────

@parcerias_metas_bp.route("/api/indicadores/orfaos", methods=["DELETE"])
@login_required
@requires_access('parcerias_metas')
def api_indicador_excluir_orfaos():
    cur = get_cursor()
    cur.execute("""
        DELETE FROM categoricas.c_dgp_indicadores
        WHERE NOT EXISTS (
            SELECT 1 FROM celebracao.celebracao_objetivos co
            WHERE categoricas.c_dgp_indicadores.id = ANY(co.indicadores_ids)
        )
        RETURNING id
    """)
    ids = [r['id'] for r in cur.fetchall()]
    get_db().commit()
    return jsonify({'excluidos': len(ids), 'ids': ids})


@parcerias_metas_bp.route("/api/meios-afericao/orfaos", methods=["DELETE"])
@login_required
@requires_access('parcerias_metas')
def api_meio_excluir_orfaos():
    cur = get_cursor()
    cur.execute("""
        DELETE FROM categoricas.c_dgp_meios_afericao
        WHERE NOT EXISTS (
            SELECT 1 FROM celebracao.celebracao_objetivos co
            WHERE categoricas.c_dgp_meios_afericao.id = ANY(co.meios_afericao_ids)
        )
        RETURNING id
    """)
    ids = [r['id'] for r in cur.fetchall()]
    get_db().commit()
    return jsonify({'excluidos': len(ids), 'ids': ids})


# ── Merge manual ──────────────────────────────────────────────────────────────

@parcerias_metas_bp.route("/api/indicadores/<int:keep_id>/fundir/<int:remove_id>", methods=["POST"])
@login_required
@requires_access('parcerias_metas')
def api_indicador_fundir(keep_id, remove_id):
    if keep_id == remove_id:
        return jsonify({'erro': 'IDs iguais'}), 400
    db  = get_db()
    cur = get_cursor()
    cur.execute("""
        UPDATE celebracao.celebracao_objetivos
        SET indicadores_ids = ARRAY_REPLACE(indicadores_ids, %s, %s)
        WHERE %s = ANY(indicadores_ids)
    """, (remove_id, keep_id, remove_id))
    cur.execute("DELETE FROM categoricas.c_dgp_indicadores WHERE id = %s", (remove_id,))
    db.commit()
    return jsonify({'ok': True})


@parcerias_metas_bp.route("/api/meios-afericao/<int:keep_id>/fundir/<int:remove_id>", methods=["POST"])
@login_required
@requires_access('parcerias_metas')
def api_meio_fundir(keep_id, remove_id):
    if keep_id == remove_id:
        return jsonify({'erro': 'IDs iguais'}), 400
    db  = get_db()
    cur = get_cursor()
    cur.execute("""
        UPDATE celebracao.celebracao_objetivos
        SET meios_afericao_ids = ARRAY_REPLACE(meios_afericao_ids, %s, %s)
        WHERE %s = ANY(meios_afericao_ids)
    """, (remove_id, keep_id, remove_id))
    cur.execute("DELETE FROM categoricas.c_dgp_meios_afericao WHERE id = %s", (remove_id,))
    db.commit()
    return jsonify({'ok': True})


# ── Reordenação drag-and-drop ─────────────────────────────────────────────────

@parcerias_metas_bp.route("/objetivo/<int:obj_id>/reordenar", methods=["PUT"])
@login_required
@requires_access('parcerias_metas')
def api_objetivo_reordenar(obj_id):
    data  = request.get_json() or {}
    ordem = data.get('ordem')
    if ordem is None:
        return jsonify({'erro': 'ordem obrigatória'}), 400
    cur = get_cursor()
    cur.execute(
        "UPDATE celebracao.celebracao_objetivos SET ordem = %s WHERE id = %s",
        (int(ordem), obj_id)
    )
    get_db().commit()
    return jsonify({'ok': True})


@parcerias_metas_bp.route("/meta/<int:meta_id>/reordenar", methods=["PUT"])
@login_required
@requires_access('parcerias_metas')
def api_meta_reordenar(meta_id):
    data  = request.get_json() or {}
    ordem = data.get('ordem')
    if ordem is None:
        return jsonify({'erro': 'ordem obrigatória'}), 400
    cur = get_cursor()
    cur.execute(
        "UPDATE celebracao.celebracao_metas SET ordem = %s WHERE id = %s",
        (int(ordem), meta_id)
    )
    get_db().commit()
    return jsonify({'ok': True})


# ── CRUD: Objetivo ────────────────────────────────────────────────────────────

@parcerias_metas_bp.route("/objetivo/criar", methods=["POST"])
@login_required
@requires_access('parcerias_metas')
def criar_objetivo():
    data    = request.get_json() or {}
    usuario = session.get('d_usuario', 'sistema')

    sei_numero = (data.get('sei_numero') or '').strip()
    objetivo   = (data.get('objetivo')   or '').strip()
    if not sei_numero or not objetivo:
        return jsonify({'erro': 'sei_numero e objetivo são obrigatórios'}), 400

    indicadores_ni = bool(data.get('indicadores_ni'))
    meios_ni       = bool(data.get('meios_ni'))

    indicadores_ids    = None if indicadores_ni else _resolve_indicadores(data.get('indicadores_textos') or [], usuario)
    meios_afericao_ids = None if meios_ni       else _resolve_meios(data.get('meios_textos') or [], usuario)

    obs_raw         = data.get('obs_indicadores') or []
    obs_indicadores = [str(o) if o else '' for o in obs_raw] or None

    cur = get_cursor()
    cur.execute(
        "SELECT COALESCE(MAX(ordem), 0) + 1 AS prox_ordem "
        "FROM celebracao.celebracao_objetivos WHERE sei_numero = %s",
        (sei_numero,)
    )
    ordem = cur.fetchone()['prox_ordem']

    cur.execute("""
        INSERT INTO celebracao.celebracao_objetivos
            (sei_numero, objetivo, indicadores_ids, indicadores_ni,
             meta_obs_indicadores, meios_afericao_ids, meios_ni, ordem, criado_por)
        VALUES (%s, %s, %s::INTEGER[], %s, %s::TEXT[], %s::INTEGER[], %s, %s, %s)
        RETURNING id
    """, (
        sei_numero, objetivo,
        indicadores_ids, indicadores_ni,
        obs_indicadores,
        meios_afericao_ids, meios_ni,
        ordem, usuario,
    ))
    novo_id = cur.fetchone()['id']
    get_db().commit()
    return jsonify({'sucesso': True, 'id': novo_id}), 201


@parcerias_metas_bp.route("/objetivo/editar/<int:obj_id>", methods=["PUT"])
@login_required
@requires_access('parcerias_metas')
def editar_objetivo(obj_id):
    data    = request.get_json() or {}
    usuario = session.get('d_usuario', 'sistema')

    objetivo = (data.get('objetivo') or '').strip()
    if not objetivo:
        return jsonify({'erro': 'objetivo é obrigatório'}), 400

    indicadores_ni = bool(data.get('indicadores_ni'))
    meios_ni       = bool(data.get('meios_ni'))

    indicadores_ids    = None if indicadores_ni else _resolve_indicadores(data.get('indicadores_textos') or [], usuario)
    meios_afericao_ids = None if meios_ni       else _resolve_meios(data.get('meios_textos') or [], usuario)

    obs_raw         = data.get('obs_indicadores') or []
    obs_indicadores = [str(o) if o else '' for o in obs_raw] or None

    cur = get_cursor()
    cur.execute("""
        UPDATE celebracao.celebracao_objetivos SET
            objetivo              = %s,
            indicadores_ids       = %s::INTEGER[],
            indicadores_ni        = %s,
            meta_obs_indicadores  = %s::TEXT[],
            meios_afericao_ids    = %s::INTEGER[],
            meios_ni              = %s,
            ordem                 = %s,
            atualizado_por        = %s,
            atualizado_em         = NOW()
        WHERE id = %s
    """, (
        objetivo,
        indicadores_ids, indicadores_ni,
        obs_indicadores,
        meios_afericao_ids, meios_ni,
        int(data.get('ordem') or 0),
        usuario,
        obj_id,
    ))
    get_db().commit()
    return jsonify({'sucesso': True})


@parcerias_metas_bp.route("/objetivo/excluir/<int:obj_id>", methods=["DELETE"])
@login_required
@requires_access('parcerias_metas')
def excluir_objetivo(obj_id):
    cur = get_cursor()
    # CASCADE exclui as metas filhas automaticamente
    cur.execute("DELETE FROM celebracao.celebracao_objetivos WHERE id = %s", (obj_id,))
    get_db().commit()
    return jsonify({'sucesso': True})


# ── CRUD: Meta ────────────────────────────────────────────────────────────────

@parcerias_metas_bp.route("/meta/criar", methods=["POST"])
@login_required
@requires_access('parcerias_metas')
def criar_meta():
    data    = request.get_json() or {}
    usuario = session.get('d_usuario', 'sistema')

    objetivo_id = data.get('objetivo_id')
    meta_titulo = (data.get('meta_titulo') or '').strip()
    if not objetivo_id or not meta_titulo:
        return jsonify({'erro': 'objetivo_id e meta_titulo são obrigatórios'}), 400

    tipos_ni      = bool(data.get('tipos_ni'))
    meta_tipo_ids = []
    if not tipos_ni:
        raw = data.get('meta_tipo_ids') or []
        if isinstance(raw, list):
            meta_tipo_ids = [int(i) for i in raw if str(i).strip().isdigit()]

    cur = get_cursor()
    cur.execute(
        "SELECT sei_numero FROM celebracao.celebracao_objetivos WHERE id = %s",
        (objetivo_id,)
    )
    obj_row = cur.fetchone()
    if not obj_row:
        return jsonify({'erro': 'Objetivo não encontrado'}), 404
    sei_numero = obj_row['sei_numero']

    cur.execute(
        "SELECT COALESCE(MAX(ordem), 0) + 1 AS prox_ordem "
        "FROM celebracao.celebracao_metas WHERE objetivo_id = %s",
        (objetivo_id,)
    )
    ordem = cur.fetchone()['prox_ordem']

    cur.execute("""
        INSERT INTO celebracao.celebracao_metas
            (objetivo_id, sei_numero, meta_titulo, meta_descricao,
             meta_tipo_ids, tipos_ni, observacoes, ordem, criado_por)
        VALUES (%s, %s, %s, %s, %s::INTEGER[], %s, %s, %s, %s)
        RETURNING id
    """, (
        objetivo_id, sei_numero,
        meta_titulo,
        data.get('meta_descricao') or None,
        meta_tipo_ids if meta_tipo_ids else None,
        tipos_ni,
        data.get('observacoes') or None,
        ordem, usuario,
    ))
    novo_id = cur.fetchone()['id']
    get_db().commit()
    return jsonify({'sucesso': True, 'id': novo_id}), 201


@parcerias_metas_bp.route("/meta/editar/<int:meta_id>", methods=["PUT"])
@login_required
@requires_access('parcerias_metas')
def editar_meta(meta_id):
    data    = request.get_json() or {}
    usuario = session.get('d_usuario', 'sistema')

    meta_titulo = (data.get('meta_titulo') or '').strip()
    if not meta_titulo:
        return jsonify({'erro': 'meta_titulo é obrigatório'}), 400

    tipos_ni      = bool(data.get('tipos_ni'))
    meta_tipo_ids = []
    if not tipos_ni:
        raw = data.get('meta_tipo_ids') or []
        if isinstance(raw, list):
            meta_tipo_ids = [int(i) for i in raw if str(i).strip().isdigit()]

    cur = get_cursor()
    cur.execute("""
        UPDATE celebracao.celebracao_metas SET
            meta_titulo    = %s,
            meta_descricao = %s,
            meta_tipo_ids  = %s::INTEGER[],
            tipos_ni       = %s,
            observacoes    = %s,
            ordem          = %s,
            atualizado_por = %s,
            atualizado_em  = NOW()
        WHERE id = %s
    """, (
        meta_titulo,
        data.get('meta_descricao') or None,
        meta_tipo_ids if meta_tipo_ids else None,
        tipos_ni,
        data.get('observacoes') or None,
        int(data.get('ordem') or 0),
        usuario,
        meta_id,
    ))
    get_db().commit()
    return jsonify({'sucesso': True})


@parcerias_metas_bp.route("/meta/excluir/<int:meta_id>", methods=["DELETE"])
@login_required
@requires_access('parcerias_metas')
def excluir_meta(meta_id):
    cur = get_cursor()
    cur.execute("DELETE FROM celebracao.celebracao_metas WHERE id = %s", (meta_id,))
    get_db().commit()
    return jsonify({'sucesso': True})


# ── Exportar CSV ──────────────────────────────────────────────────────────────

@parcerias_metas_bp.route("/exportar-csv", methods=["GET"])
@login_required
@requires_access('parcerias_metas')
def exportar_csv():
    cur = get_cursor()
    filtro_sei   = request.args.get('filtro_sei',   '').strip()
    filtro_ind   = request.args.get('filtro_ind',   '').strip()
    filtro_meio  = request.args.get('filtro_meio',  '').strip()
    filtro_texto = request.args.get('filtro_texto', '').strip()

    where_clause = ""
    conditions = []
    params = []
    if filtro_sei:
        conditions.append("co.sei_numero ILIKE %s")
        params.append(f"%{filtro_sei}%")
    if filtro_ind:
        conditions.append("""EXISTS (
            SELECT 1 FROM categoricas.c_dgp_indicadores _ind
            WHERE _ind.id = ANY(co.indicadores_ids)
              AND LOWER(_ind.indicador) = LOWER(%s)
        )""")
        params.append(filtro_ind)
    if filtro_meio:
        conditions.append("""EXISTS (
            SELECT 1 FROM categoricas.c_dgp_meios_afericao _ma
            WHERE _ma.id = ANY(co.meios_afericao_ids)
              AND LOWER(_ma.meios_afericao) = LOWER(%s)
        )""")
        params.append(filtro_meio)
    if filtro_texto:
        conditions.append("""(
            co.objetivo ILIKE %s
            OR EXISTS (
                SELECT 1 FROM celebracao.celebracao_metas _cm2
                WHERE _cm2.objetivo_id = co.id
                  AND (_cm2.meta_titulo ILIKE %s OR _cm2.meta_descricao ILIKE %s)
            )
        )""")
        params.extend([f"%{filtro_texto}%"] * 3)
    if conditions:
        where_clause = "WHERE " + " AND ".join(conditions)

    cur.execute(f"""
        SELECT
            co.sei_numero,
            COALESCE(p.numero_termo, cp.numero_termo, '-')    AS numero_termo,
            COALESCE(p.osc, cp.osc, '-')                      AS osc,
            co.objetivo,
            COALESCE(
                (SELECT STRING_AGG(ind.indicador, ' | ' ORDER BY ind.indicador)
                 FROM categoricas.c_dgp_indicadores ind
                 WHERE ind.id = ANY(co.indicadores_ids)),
                ''
            ) AS indicadores,
            co.meta_obs_indicadores,
            COALESCE(
                (SELECT STRING_AGG(ma.meios_afericao, ' | ' ORDER BY ma.meios_afericao)
                 FROM categoricas.c_dgp_meios_afericao ma
                 WHERE ma.id = ANY(co.meios_afericao_ids)),
                ''
            ) AS meios_afericao,
            cm.ordem   AS meta_ordem,
            cm.meta_titulo,
            cm.meta_descricao,
            COALESCE(
                (SELECT STRING_AGG(mt.meta_tipo || ' (' || COALESCE(mt.tipo_classificacao,'') || ')', ' | '
                                   ORDER BY mt.tipo_classificacao, mt.meta_tipo)
                 FROM categoricas.c_dgp_meta_tipos mt
                 WHERE mt.id = ANY(cm.meta_tipo_ids)),
                ''
            ) AS tipos_labels,
            cm.observacoes,
            cm.criado_por  AS meta_criado_por,
            cm.criado_em   AS meta_criado_em
        FROM celebracao.celebracao_objetivos co
        LEFT JOIN celebracao.celebracao_metas cm ON cm.objetivo_id = co.id
        LEFT JOIN (SELECT sei_celeb, MAX(numero_termo) AS numero_termo, MAX(osc) AS osc
                   FROM public.parcerias GROUP BY sei_celeb) p ON p.sei_celeb = co.sei_numero
        LEFT JOIN (SELECT sei_celeb, MAX(numero_termo) AS numero_termo, MAX(osc) AS osc
                   FROM celebracao.celebracao_parcerias GROUP BY sei_celeb) cp ON cp.sei_celeb = co.sei_numero
        {where_clause}
        ORDER BY co.sei_numero, co.ordem, co.id, cm.ordem, cm.id
    """, params)
    rows = cur.fetchall()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_ALL)
    writer.writerow([
        'SEI', 'Nº Termo', 'OSC',
        'Objetivo', 'Indicadores', 'Obs. Indicadores', 'Meios de Aferição',
        'Ord. Meta', 'Título da Meta', 'Descrição', 'Tipos', 'Observações',
        'Criado Por', 'Criado Em',
    ])
    for r in rows:
        obs_str = ' | '.join(r['meta_obs_indicadores'] or []) if r['meta_obs_indicadores'] else ''
        writer.writerow([
            r['sei_numero'], r['numero_termo'], r['osc'],
            r['objetivo'],
            r['indicadores'] or '',
            obs_str,
            r['meios_afericao'] or '',
            r['meta_ordem'] if r['meta_ordem'] is not None else '',
            r['meta_titulo'] or '',
            r['meta_descricao'] or '',
            r['tipos_labels'] or '',
            r['observacoes'] or '',
            r['meta_criado_por'] or '',
            r['meta_criado_em'].strftime('%d/%m/%Y %H:%M') if r['meta_criado_em'] else '',
        ])

    output.seek(0)
    filename = f"quadro_metas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return Response(
        '\ufeff' + output.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


# ── Exportar Excel ────────────────────────────────────────────────────────────

@parcerias_metas_bp.route("/exportar-excel", methods=["GET"])
@login_required
@requires_access('parcerias_metas')
def exportar_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    cur = get_cursor()
    filtro_sei   = request.args.get('filtro_sei',   '').strip()
    filtro_ind   = request.args.get('filtro_ind',   '').strip()
    filtro_meio  = request.args.get('filtro_meio',  '').strip()
    filtro_texto = request.args.get('filtro_texto', '').strip()

    where_clause = ""
    conditions = []
    params = []
    if filtro_sei:
        conditions.append("co.sei_numero ILIKE %s")
        params.append(f"%{filtro_sei}%")
    if filtro_ind:
        conditions.append("""EXISTS (
            SELECT 1 FROM categoricas.c_dgp_indicadores _ind
            WHERE _ind.id = ANY(co.indicadores_ids)
              AND LOWER(_ind.indicador) = LOWER(%s)
        )""")
        params.append(filtro_ind)
    if filtro_meio:
        conditions.append("""EXISTS (
            SELECT 1 FROM categoricas.c_dgp_meios_afericao _ma
            WHERE _ma.id = ANY(co.meios_afericao_ids)
              AND LOWER(_ma.meios_afericao) = LOWER(%s)
        )""")
        params.append(filtro_meio)
    if filtro_texto:
        conditions.append("""(
            co.objetivo ILIKE %s
            OR EXISTS (
                SELECT 1 FROM celebracao.celebracao_metas _cm3
                WHERE _cm3.objetivo_id = co.id
                  AND (_cm3.meta_titulo ILIKE %s OR _cm3.meta_descricao ILIKE %s)
            )
        )""")
        params.extend([f"%{filtro_texto}%"] * 3)
    if conditions:
        where_clause = "WHERE " + " AND ".join(conditions)

    cur.execute(f"""
        SELECT
            co.sei_numero,
            COALESCE(p.numero_termo, cp.numero_termo, '-')    AS numero_termo,
            COALESCE(p.osc, cp.osc, '-')                      AS osc,
            cp.status                                         AS status,
            co.objetivo,
            COALESCE(
                (SELECT STRING_AGG(ind.indicador, ' | ' ORDER BY ind.indicador)
                 FROM categoricas.c_dgp_indicadores ind
                 WHERE ind.id = ANY(co.indicadores_ids)), ''
            ) AS indicadores,
            co.meta_obs_indicadores,
            COALESCE(
                (SELECT STRING_AGG(ma.meios_afericao, ' | ' ORDER BY ma.meios_afericao)
                 FROM categoricas.c_dgp_meios_afericao ma
                 WHERE ma.id = ANY(co.meios_afericao_ids)), ''
            ) AS meios_afericao,
            cm.ordem   AS meta_ordem,
            cm.meta_titulo,
            cm.meta_descricao,
            COALESCE(
                (SELECT STRING_AGG(mt.meta_tipo || ' (' || COALESCE(mt.tipo_classificacao,'') || ')', ' | '
                                   ORDER BY mt.tipo_classificacao, mt.meta_tipo)
                 FROM categoricas.c_dgp_meta_tipos mt
                 WHERE mt.id = ANY(cm.meta_tipo_ids)), ''
            ) AS tipos_labels,
            cm.observacoes,
            cm.criado_por  AS meta_criado_por,
            cm.criado_em   AS meta_criado_em
        FROM celebracao.celebracao_objetivos co
        LEFT JOIN celebracao.celebracao_metas cm ON cm.objetivo_id = co.id
        LEFT JOIN (SELECT sei_celeb, MAX(numero_termo) AS numero_termo, MAX(osc) AS osc
                   FROM public.parcerias GROUP BY sei_celeb) p ON p.sei_celeb = co.sei_numero
        LEFT JOIN (SELECT sei_celeb, MAX(numero_termo) AS numero_termo, MAX(osc) AS osc, MAX(status) AS status
                   FROM celebracao.celebracao_parcerias GROUP BY sei_celeb) cp ON cp.sei_celeb = co.sei_numero
        {where_clause}
        ORDER BY co.sei_numero, co.ordem, co.id, cm.ordem, cm.id
    """, params)
    rows = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quadro de Metas"

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="4c1d95")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    headers = [
        'SEI', 'Nº Termo', 'OSC', 'Status',
        'Objetivo', 'Indicadores', 'Obs. Indicadores', 'Meios de Aferição',
        'Ord. Meta', 'Título da Meta', 'Descrição', 'Tipos', 'Observações',
        'Criado Por', 'Criado Em',
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align
    ws.row_dimensions[1].height = 30

    for r in rows:
        obs_str = ' | '.join(r['meta_obs_indicadores'] or []) if r['meta_obs_indicadores'] else ''
        ws.append([
            r['sei_numero'], r['numero_termo'], r['osc'],
            r['status'] or '',
            r['objetivo'],
            r['indicadores'] or '',
            obs_str,
            r['meios_afericao'] or '',
            r['meta_ordem'] if r['meta_ordem'] is not None else '',
            r['meta_titulo'] or '',
            r['meta_descricao'] or '',
            r['tipos_labels'] or '',
            r['observacoes'] or '',
            r['meta_criado_por'] or '',
            r['meta_criado_em'].strftime('%d/%m/%Y %H:%M') if r['meta_criado_em'] else '',
        ])

    # Ajustar larguras
    col_widths = [18, 18, 35, 22, 45, 45, 35, 45, 8, 35, 55, 35, 35, 20, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"quadro_metas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )

