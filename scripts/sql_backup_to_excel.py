"""
sql_backup_to_excel.py
======================
Converte um backup PostgreSQL (.sql com COPY ... FROM stdin) em arquivos Excel
organizados por schema — um arquivo .xlsx por schema, uma aba por tabela.

Uso:
    python scripts/sql_backup_to_excel.py                           # usa o backup mais recente em backups/
    python scripts/sql_backup_to_excel.py caminho/para/backup.sql  # arquivo específico
    python scripts/sql_backup_to_excel.py --csv                    # exporta CSVs em vez de XLSX

Saída: pasta  backups/exportado_<nome_backup>/
"""

import sys
import os
import csv
import re
from pathlib import Path
from collections import defaultdict

# Detecta se openpyxl está disponível
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------
SCHEMAS_IGNORADOS = {"auth", "storage", "realtime", "extensions", "graphql",
                     "graphql_public", "pgbouncer", "pgsodium", "vault",
                     "supabase_functions", "_realtime"}

NULL_MARKER = "\\N"   # marcador de NULL no formato COPY do PostgreSQL

# Limite para evitar abas gigantes no Excel (linhas por aba)
MAX_ROWS_POR_ABA = 500_000


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

def _decode_copy_value(val: str):
    """Converte escapes do formato COPY do PostgreSQL para Python."""
    if val == NULL_MARKER:
        return None
    # Desescapar sequências padrão do pg COPY
    val = val.replace("\\\\", "\x00BKSL\x00")  # preserva \\ temporariamente
    val = val.replace("\\t", "\t")
    val = val.replace("\\n", "\n")
    val = val.replace("\\r", "\r")
    val = val.replace("\x00BKSL\x00", "\\")
    return val


def _parse_copy_header(line: str):
    """
    Extrai (schema, tabela, [colunas]) de uma linha COPY.
    Exemplos aceitos:
        COPY public.parcerias (id, numero_termo, nome) FROM stdin;
        COPY analises_pc.conc_extrato (id, data, valor) FROM stdin;
    """
    m = re.match(
        r"COPY\s+([\w]+)\.([\w]+)\s*\(([^)]+)\)\s*FROM\s+stdin;",
        line, re.IGNORECASE
    )
    if not m:
        return None, None, []
    schema = m.group(1)
    tabela = m.group(2)
    colunas = [c.strip() for c in m.group(3).split(",")]
    return schema, tabela, colunas


def parse_sql_dump(sql_path: str, schemas_ignorados=None):
    """
    Lê o dump SQL e retorna:
        dados[schema][tabela] = {"colunas": [...], "linhas": [[...], ...]}
    """
    if schemas_ignorados is None:
        schemas_ignorados = SCHEMAS_IGNORADOS

    dados = defaultdict(dict)
    schema_atual = None
    tabela_atual = None
    colunas_atuais = []
    em_copy = False
    linhas_atuais = []

    total_linhas_dados = 0
    tabelas_lidas = 0

    print(f"\n📂 Lendo: {sql_path}")

    with open(sql_path, "r", encoding="utf-8", errors="replace") as f:
        for i, raw_line in enumerate(f):
            line = raw_line.rstrip("\n").rstrip("\r")

            if not em_copy:
                if line.upper().startswith("COPY "):
                    schema_atual, tabela_atual, colunas_atuais = _parse_copy_header(line)
                    if schema_atual and schema_atual not in schemas_ignorados:
                        em_copy = True
                        linhas_atuais = []
            else:
                # Terminador do bloco COPY
                if line == "\\.":
                    dados[schema_atual][tabela_atual] = {
                        "colunas": colunas_atuais,
                        "linhas": linhas_atuais,
                    }
                    total_linhas_dados += len(linhas_atuais)
                    tabelas_lidas += 1
                    if tabelas_lidas % 10 == 0:
                        print(f"   ... {tabelas_lidas} tabelas lidas, "
                              f"{total_linhas_dados:,} linhas de dados")
                    em_copy = False
                    schema_atual = None
                else:
                    campos = [_decode_copy_value(v) for v in line.split("\t")]
                    linhas_atuais.append(campos)

    print(f"\n✅ Leitura concluída: {tabelas_lidas} tabelas, "
          f"{total_linhas_dados:,} linhas de dados")
    return dados


# ---------------------------------------------------------------------------
# Exportação para XLSX
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79",
                          fill_type="solid") if XLSX_AVAILABLE else None
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10) if XLSX_AVAILABLE else None
HEADER_ALIGN = Alignment(horizontal="center",
                         vertical="center") if XLSX_AVAILABLE else None


def _escrever_aba(ws, colunas, linhas):
    """Preenche uma aba do workbook com cabeçalho estilizado e dados."""
    # Cabeçalho
    for col_idx, nome_col in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=nome_col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # Dados (limitado a MAX_ROWS_POR_ABA para evitar crashar o Excel)
    for row_idx, linha in enumerate(linhas[:MAX_ROWS_POR_ABA], start=2):
        for col_idx, valor in enumerate(linha, start=1):
            # Converte None para string vazia no Excel
            ws.cell(row=row_idx, column=col_idx, value=valor if valor is not None else "")

    # Ajusta largura automaticamente (máx 50 chars)
    for col_idx, nome_col in enumerate(colunas, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        max_len = len(nome_col)
        for linha in linhas[:200]:  # amostra para performance
            if col_idx - 1 < len(linha) and linha[col_idx - 1] is not None:
                max_len = max(max_len, len(str(linha[col_idx - 1])))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    # Congela a primeira linha
    ws.freeze_panes = "A2"


def exportar_xlsx(dados: dict, pasta_saida: Path):
    """Cria um arquivo .xlsx por schema."""
    if not XLSX_AVAILABLE:
        print("⚠️  openpyxl não encontrado. Instale com: pip install openpyxl")
        return

    pasta_saida.mkdir(parents=True, exist_ok=True)

    for schema, tabelas in sorted(dados.items()):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove aba vazia padrão

        print(f"\n📊 Schema '{schema}' — {len(tabelas)} tabelas:")

        for tabela, info in sorted(tabelas.items()):
            n_linhas = len(info["linhas"])
            print(f"   • {tabela:45s} {n_linhas:>8,} linhas")

            # Nome da aba limitado a 31 chars (limite do Excel)
            nome_aba = tabela[:31]
            ws = wb.create_sheet(title=nome_aba)

            # Metadado na célula A1 se tabela foi truncada
            if n_linhas > MAX_ROWS_POR_ABA:
                ws.sheet_properties.tabColor = "FF0000"  # aba vermelha = truncada

            _escrever_aba(ws, info["colunas"], info["linhas"])

            if n_linhas > MAX_ROWS_POR_ABA:
                print(f"     ⚠️  truncada em {MAX_ROWS_POR_ABA:,} linhas "
                      f"(total: {n_linhas:,})")

        arquivo = pasta_saida / f"{schema}.xlsx"
        wb.save(arquivo)
        print(f"   💾 Salvo: {arquivo}")


# ---------------------------------------------------------------------------
# Exportação para CSV
# ---------------------------------------------------------------------------

def exportar_csv(dados: dict, pasta_saida: Path):
    """Cria um CSV por tabela, em subpastas por schema."""
    for schema, tabelas in sorted(dados.items()):
        pasta_schema = pasta_saida / schema
        pasta_schema.mkdir(parents=True, exist_ok=True)

        print(f"\n📁 Schema '{schema}' — {len(tabelas)} tabelas:")

        for tabela, info in sorted(tabelas.items()):
            arquivo = pasta_schema / f"{tabela}.csv"
            n_linhas = len(info["linhas"])
            print(f"   • {tabela:45s} {n_linhas:>8,} linhas  → {arquivo.name}")

            with open(arquivo, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
                writer.writerow(info["colunas"])
                for linha in info["linhas"]:
                    # None → string vazia para CSV
                    writer.writerow([v if v is not None else "" for v in linha])

    print(f"\n💾 CSVs salvos em: {pasta_saida}")


# ---------------------------------------------------------------------------
# Utilitário: encontrar backup mais recente
# ---------------------------------------------------------------------------

def encontrar_backup_mais_recente(pasta_backups: Path) -> Path | None:
    arquivos = sorted(pasta_backups.glob("backup_*.sql"), reverse=True)
    return arquivos[0] if arquivos else None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = sys.argv[1:]
    usar_csv = "--csv" in args
    args = [a for a in args if a != "--csv"]

    # Resolve o arquivo de backup
    if args:
        sql_path = Path(args[0])
    else:
        pasta_backups = Path(__file__).parent.parent / "backups"
        sql_path = encontrar_backup_mais_recente(pasta_backups)
        if sql_path is None:
            print("❌ Nenhum arquivo backup_*.sql encontrado em backups/")
            sys.exit(1)
        print(f"🔍 Backup mais recente: {sql_path.name}")

    if not sql_path.exists():
        print(f"❌ Arquivo não encontrado: {sql_path}")
        sys.exit(1)

    # Pasta de saída
    nome_base = sql_path.stem  # ex: backup_faf_20260320_095637
    pasta_saida = sql_path.parent / f"exportado_{nome_base}"

    if not usar_csv and not XLSX_AVAILABLE:
        print("⚠️  openpyxl não instalado. Mudando para modo CSV.")
        print("   Para XLSX, execute: pip install openpyxl\n")
        usar_csv = True

    # Parse
    dados = parse_sql_dump(str(sql_path))

    if not dados:
        print("⚠️  Nenhum dado encontrado no backup.")
        sys.exit(0)

    # Exporta
    modo = "CSV" if usar_csv else "XLSX"
    print(f"\n{'='*60}")
    print(f"Exportando {len(dados)} schemas no formato {modo}...")
    print(f"Destino: {pasta_saida}")
    print("="*60)

    if usar_csv:
        exportar_csv(dados, pasta_saida)
    else:
        exportar_xlsx(dados, pasta_saida)

    print(f"\n✅ Exportação concluída! Pasta: {pasta_saida}")


if __name__ == "__main__":
    main()
